import openpyxl
from pathlib import Path
files = [
 Path(r"c:/Users/ClaudiaAlvarezDiaz/Documents/EJIE/Repo OSAKIDETZA/Osakidetza/Context/Notas reuniones/Dudas y preguntas (4).xlsx"),
 Path(r"c:/Users/ClaudiaAlvarezDiaz/Documents/EJIE/Repo OSAKIDETZA/Osakidetza/Estimación de tareas/OSAKIDETZA - Prescripciones Ortoprotésicas - Tareas (1) (1).xlsx"),
]
for f in files:
    print("\n"+"="*80)
    print('FILE:', f)
    try:
        wb = openpyxl.load_workbook(f, data_only=True)
    except Exception as e:
        print('ERROR opening', e)
        continue
    for ws in wb.worksheets:
        print(f'-- SHEET: {ws.title}')
        max_r = min(ws.max_row, 80)
        max_c = min(ws.max_column, 14)
        for r in range(1, max_r+1):
            vals=[]
            for c in range(1, max_c+1):
                v=ws.cell(r,c).value
                if v is not None and str(v).strip()!='':
                    vals.append(f'C{c}:{v}')
            if vals:
                print(f'R{r} | ' + ' ; '.join(vals))
